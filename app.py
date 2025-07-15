import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import json
import os
import time
from datetime import datetime
import re
import io
import zipfile
import hashlib
from typing import Dict, List, Any, Tuple, Optional
from openai import OpenAI
import requests
from docx import Document
from docx.shared import Inches
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.oxml.shared import OxmlElement, qn
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import Color, red, orange, yellow, lightgrey, black
from reportlab.lib.units import inch
from reportlab.platypus.flowables import KeepTogether
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
import PyPDF2
import pdfplumber
from PIL import Image

# Configuration
MAX_CHARS_PER_CHUNK = 100000
CHUNK_DELAY = 0.5
MAX_RETRIES = 3
VIOLATION_RULES = {
    "violence": "Excessive violence must be toned down or justified by plot",
    "language": "Offensive language should be minimized",
    "nudity": "Nudity requires careful consideration and context",
    "drug_use": "Drug use depictions must be handled responsibly",
    "copyright": "Potential copyright issues must be addressed"
}

# Page configuration
st.set_page_config(
    page_title="Film Production Master App",
    page_icon="🎬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(90deg, #ff6b6b, #4ecdc4);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    .violation-critical {
        background-color: #ffebee;
        border-left: 5px solid #f44336;
        padding: 10px;
        margin: 5px 0;
    }
    .violation-high {
        background-color: #fff3e0;
        border-left: 5px solid #ff9800;
        padding: 10px;
        margin: 5px 0;
    }
    .violation-medium {
        background-color: #fffde7;
        border-left: 5px solid #ffeb3b;
        padding: 10px;
        margin: 5px 0;
    }
    .violation-low {
        background-color: #f3e5f5;
        border-left: 5px solid #9c27b0;
        padding: 10px;
        margin: 5px 0;
    }
    .scene-header {
        background-color: #e3f2fd;
        padding: 8px;
        border-radius: 5px;
        font-weight: bold;
        margin: 10px 0;
    }
    .comment-box {
        background-color: #f5f5f5;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
    .resolved-comment {
        background-color: #e8f5e8;
        border-left: 3px solid #4caf50;
    }
    .user-info {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #007bff;
    }
    .tab-container {
        padding: 15px;
        border-radius: 10px;
        background-color: #f9f9f9;
        margin-bottom: 20px;
    }
</style>
""", unsafe_allow_html=True)

# Authentication functions
def check_email_domain(email: str) -> bool:
    """Check if email belongs to authorized domain"""
    authorized_domains = ['@hoichoi.tv', '@gmail.com', '@example.com']
    return any(email.lower().strip().endswith(domain) for domain in authorized_domains)

def hash_password(password: str) -> str:
    """Simple password hashing"""
    return hashlib.sha256(password.encode()).hexdigest()

def authenticate_user():
    """Handle user authentication"""
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        st.markdown("""
        <div class="main-header">
            <h1>🎬 Film Production Master App</h1>
            <h3>Standards & Practices + Production Design</h3>
        </div>
        """, unsafe_allow_html=True)
        
        with st.container():
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.subheader("🔐 Login Required")
                email = st.text_input("Email Address", placeholder="yourname@hoichoi.tv")
                password = st.text_input("Password", type="password")
                
                if st.button("Login", type="primary", use_container_width=True):
                    if check_email_domain(email):
                        st.session_state.authenticated = True
                        st.session_state.user_email = email
                        st.session_state.user_name = email.split('@')[0].replace('.', ' ').title()
                        st.session_state.is_admin = email in ['admin@hoichoi.tv', 'sp@hoichoi.tv']
                        st.rerun()
                    else:
                        st.error("❌ Access denied. Please use an authorized email address.")
        
        return False
    return True

# API Key Management
def get_deepseek_api_key():
    """Get DeepSeek API key from Streamlit secrets or user input"""
    try:
        if hasattr(st.session_state, 'temp_deepseek_key') and st.session_state.temp_deepseek_key:
            return st.session_state.temp_deepseek_key
        return st.secrets.get("DEEPSEEK_API_KEY", None)
    except:
        return None

# Film Script Processing Class
class FilmScriptProcessor:
    def __init__(self, api_key: str):
        """Initialize the processor with API key"""
        self.client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com/v1")
        self.snp_system_prompt = """You are a senior film standards and practices reviewer. Analyze the script for compliance with content guidelines."""
        self.production_system_prompt = """You are a senior film production coordinator. Extract production elements including locations, scenes, and props."""
    
    def extract_text_from_pdf(self, file_data: bytes) -> str:
        """Extract text from PDF bytes"""
        try:
            text = ""
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_data))
            for i, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text()
                text += page_text + "\n"
            return text
        except Exception as e:
            raise Exception(f"Error reading PDF: {str(e)}")

    def extract_text_from_docx(self, file_data: bytes) -> str:
        """Extract text from DOCX bytes"""
        try:
            doc = Document(io.BytesIO(file_data))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            raise Exception(f"Error reading DOCX: {str(e)}")

    def extract_text_from_file(self, file_data: bytes, filename: str) -> str:
        """Extract text from file based on extension"""
        file_extension = os.path.splitext(filename)[1].lower()
        
        if file_extension == '.pdf':
            return self.extract_text_from_pdf(file_data)
        elif file_extension in ['.docx', '.doc']:
            return self.extract_text_from_docx(file_data)
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")

    def analyze_for_standards(self, text: str) -> Dict[str, Any]:
        """Analyze script for standards and practices violations"""
        try:
            response = self.client.chat.completions.create(
                model="deepseek-chat",
                messages=[
                    {"role": "system", "content": self.snp_system_prompt},
                    {"role": "user", "content": f"""Analyze this script for content violations. Check for:
                    - Excessive violence
                    - Offensive language
                    - Nudity/sexual content
                    - Drug use
                    - Copyright issues
                    Return JSON with violations found:
                    {{
                        "violations": [
                            {{
                                "violationType": "violence",
                                "severity": "high",
                                "violationText": "excerpt from script",
                                "explanation": "why it violates standards",
                                "suggestedAction": "how to fix",
                                "pageNumber": 1
                            }}
                        ],
                        "summary": {{
                            "totalPages": 10,
                            "successRate": "100%"
                        }}
                    }}
                    Script: {text[:10000]}..."""  # Limiting to first 10k chars for demo
                ],
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            
            result = json.loads(response.choices[0].message.content)
            return result
        except Exception as e:
            return {"error": str(e)}

    def analyze_for_production(self, text: str) -> Dict[str, Any]:
        """Analyze script for production elements"""
        try:
            response = self.client.chat.completions.create(
                model="deepseek-chat",
                messages=[
                    {"role": "system", "content": self.production_system_prompt},
                    {"role": "user", "content": f"""Extract production elements from this script. Return JSON with:
                    {{
                        "location_breakdown": [
                            {{
                                "location_name": "Office",
                                "scenes_in_location": [
                                    {{
                                        "scene_number": 1,
                                        "scene_heading": "INT. OFFICE - DAY",
                                        "time_of_day": "DAY",
                                        "brief_description": "Character enters office",
                                        "props_in_scene": ["desk", "chair"]
                                    }}
                                ]
                            }}
                        ],
                        "unique_props": ["desk", "chair"]
                    }}
                    Script: {text[:10000]}..."""  # Limiting to first 10k chars for demo
                ],
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            
            result = json.loads(response.choices[0].message.content)
            return result
        except Exception as e:
            return {"error": str(e)}

# Report Generation Functions
def generate_snp_report(violations: List[Dict]) -> Tuple[bytes, bytes]:
    """Generate reports for standards and practices"""
    # Generate XLSX
    df = pd.DataFrame(violations)
    xlsx_buffer = io.BytesIO()
    
    with pd.ExcelWriter(xlsx_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Violations', index=False)
        
        # Summary
        summary = pd.DataFrame({
            'Metric': ['Total', 'Critical', 'High', 'Medium', 'Low'],
            'Count': [
                len(violations),
                len([v for v in violations if v.get('severity') == 'critical']),
                len([v for v in violations if v.get('severity') == 'high']),
                len([v for v in violations if v.get('severity') == 'medium']),
                len([v for v in violations if v.get('severity') == 'low'])
            ]
        })
        summary.to_excel(writer, sheet_name='Summary', index=False)
    
    xlsx_buffer.seek(0)
    
    # Generate PDF content
    pdf_content = f"""STANDARDS & PRACTICES REPORT
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SUMMARY:
Total Violations: {len(violations)}
Critical: {len([v for v in violations if v.get('severity') == 'critical'])}
High: {len([v for v in violations if v.get('severity') == 'high'])}

VIOLATIONS:
"""
    
    for i, v in enumerate(violations[:10], 1):
        pdf_content += f"\n{i}. {v.get('violationType', 'Unknown')} (Page {v.get('pageNumber', 'N/A')})\n"
        pdf_content += f"   Text: {v.get('violationText', 'N/A')[:100]}...\n"
        pdf_content += f"   Action: {v.get('suggestedAction', 'N/A')}\n"
    
    return xlsx_buffer.getvalue(), pdf_content.encode('utf-8')

def generate_production_report(results: Dict[str, Any]) -> bytes:
    """Generate Excel report for production breakdown"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Summary sheet
    summary_sheet = wb.create_sheet("SUMMARY")
    summary_sheet['A1'] = "PRODUCTION BREAKDOWN"
    summary_sheet['A1'].font = Font(bold=True, size=16)
    
    # Add data
    total_locations = len(results.get('location_breakdown', []))
    total_scenes = sum(len(loc.get('scenes_in_location', [])) for loc in results.get('location_breakdown', []))
    total_props = len(results.get('unique_props', []))
    
    summary_sheet['A3'] = "Total Locations:"
    summary_sheet['B3'] = total_locations
    summary_sheet['A4'] = "Total Scenes:"
    summary_sheet['B4'] = total_scenes
    summary_sheet['A5'] = "Total Props:"
    summary_sheet['B5'] = total_props
    
    # Location sheets
    for location in results.get('location_breakdown', []):
        sheet = wb.create_sheet(location['location_name'][:30])
        sheet.append(["Scene", "Heading", "Time", "Description", "Props"])
        
        for scene in location['scenes_in_location']:
            sheet.append([
                scene['scene_number'],
                scene['scene_heading'],
                scene['time_of_day'],
                scene['brief_description'],
                ", ".join(scene['props_in_scene'])
            ])
    
    # Props sheet
    props_sheet = wb.create_sheet("PROPS")
    props_sheet.append(["Prop", "Locations Used"])
    
    for prop in results.get('unique_props', []):
        locations = []
        for location in results.get('location_breakdown', []):
            for scene in location['scenes_in_location']:
                if prop in scene['props_in_scene'] and location['location_name'] not in locations:
                    locations.append(location['location_name'])
        
        props_sheet.append([prop, ", ".join(locations)])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# Main App Tabs
def tab_upload_script():
    """Main upload and processing tab"""
    st.header("📤 Upload Script")
    st.markdown("Upload your script for both Standards & Practices review and Production Design analysis.")
    
    uploaded_file = st.file_uploader(
        "Choose a script file (PDF or DOCX)",
        type=['pdf', 'docx'],
        help="Upload your screenplay for analysis"
    )
    
    if uploaded_file:
        st.success(f"✅ File uploaded: {uploaded_file.name}")
        
        deepseek_key = get_deepseek_api_key()
        if not deepseek_key:
            st.error("❌ DeepSeek API key required for processing")
            return
        
        processor = FilmScriptProcessor(deepseek_key)
        
        if st.button("🔍 Analyze Script", type="primary"):
            with st.spinner("Processing script..."):
                try:
                    # Extract text
                    text = processor.extract_text_from_file(uploaded_file.getvalue(), uploaded_file.name)
                    
                    # Process for both standards and production
                    standards_result = processor.analyze_for_standards(text)
                    production_result = processor.analyze_for_production(text)
                    
                    # Save to session state
                    st.session_state.standards_results = standards_result
                    st.session_state.production_results = production_result
                    st.session_state.current_filename = uploaded_file.name
                    
                    st.success("🎉 Analysis complete!")
                    
                except Exception as e:
                    st.error(f"❌ Processing error: {str(e)}")

def tab_standards_results():
    """Display standards and practices results"""
    if 'standards_results' not in st.session_state:
        st.info("ℹ️ Upload and analyze a script to see Standards & Practices results")
        return
    
    st.header("📝 Standards & Practices Review")
    
    results = st.session_state.standards_results
    if 'error' in results:
        st.error(f"❌ Error in analysis: {results['error']}")
        return
    
    violations = results.get('violations', [])
    
    # Summary stats
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Violations", len(violations))
    with col2:
        st.metric("Critical", len([v for v in violations if v.get('severity') == 'critical']))
    with col3:
        st.metric("High", len([v for v in violations if v.get('severity') == 'high']))
    
    # Violation breakdown
    st.subheader("Violation Breakdown")
    for violation in violations:
        severity = violation.get('severity', 'low')
        violation_class = f"violation-{severity}"
        
        with st.container():
            st.markdown(f"""
            <div class="{violation_class}">
                <h4>🚨 {violation.get('violationType', 'Unknown')} (Page {violation.get('pageNumber', 'N/A')})</h4>
                <p><strong>Severity:</strong> {severity.upper()}</p>
                <p><strong>Text:</strong> "{violation.get('violationText', 'N/A')}"</p>
                <p><strong>Issue:</strong> {violation.get('explanation', 'N/A')}</p>
                <p><strong>Action:</strong> {violation.get('suggestedAction', 'N/A')}</p>
            </div>
            """, unsafe_allow_html=True)
    
    # Download reports
    st.subheader("📥 Download Reports")
    if st.button("Generate S&P Reports"):
        xlsx_data, pdf_data = generate_snp_report(violations)
        
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📊 Download Excel Report",
                data=xlsx_data,
                file_name=f"{st.session_state.current_filename}_snp_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col2:
            st.download_button(
                label="📄 Download PDF Report",
                data=pdf_data,
                file_name=f"{st.session_state.current_filename}_snp_report.pdf",
                mime="application/pdf"
            )

def tab_production_results():
    """Display production design results"""
    if 'production_results' not in st.session_state:
        st.info("ℹ️ Upload and analyze a script to see Production Design results")
        return
    
    st.header("🎬 Production Design Breakdown")
    
    results = st.session_state.production_results
    if 'error' in results:
        st.error(f"❌ Error in analysis: {results['error']}")
        return
    
    # Summary stats
    total_locations = len(results.get('location_breakdown', []))
    total_scenes = sum(len(loc.get('scenes_in_location', [])) for loc in results.get('location_breakdown', []))
    total_props = len(results.get('unique_props', []))
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Locations", total_locations)
    with col2:
        st.metric("Total Scenes", total_scenes)
    with col3:
        st.metric("Total Props", total_props)
    
    # Location breakdown
    st.subheader("Location Breakdown")
    for location in results.get('location_breakdown', []):
        with st.expander(f"📍 {location['location_name']} ({len(location['scenes_in_location'])} scenes)"):
            for scene in location['scenes_in_location']:
                st.write(f"**Scene {scene.get('scene_number', 'N/A')}**: {scene.get('scene_heading', 'N/A')}")
                st.write(f"*Time*: {scene.get('time_of_day', 'N/A')}")
                st.write(f"*Props*: {', '.join(scene.get('props_in_scene', []))}")
                st.divider()
    
    # Props list
    st.subheader("Master Props List")
    if results.get('unique_props'):
        props_df = pd.DataFrame({
            'Prop': results['unique_props'],
            'Index': range(1, len(results['unique_props']) + 1)
        })
        st.dataframe(props_df, use_container_width=True)
    
    # Download report
    st.subheader("📥 Download Report")
    if st.button("Generate Production Report"):
        excel_data = generate_production_report(results)
        
        st.download_button(
            label="📊 Download Excel Report",
            data=excel_data,
            file_name=f"{st.session_state.current_filename}_production_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# Main Application
def main():
    # Authentication check
    if not authenticate_user():
        return
    
    # Sidebar
    with st.sidebar:
        st.markdown(f"""
        <div class="user-info">
            <h3>👤 User Information</h3>
            <p><b>Name:</b> {st.session_state.get('user_name', 'Unknown')}</p>
            <p><b>Email:</b> {st.session_state.get('user_email', 'unknown')}</p>
            <p><b>Role:</b> {'Admin' if st.session_state.get('is_admin', False) else 'User'}</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        
        # API Key Configuration
        st.subheader("🔑 API Configuration")
        deepseek_key = get_deepseek_api_key()
        if deepseek_key:
            st.success("✅ DeepSeek API Key: Configured")
        else:
            st.warning("⚠️ DeepSeek API Key: Not configured")
            deepseek_input = st.text_input(
                "Enter DeepSeek API Key", 
                type="password", 
                help="Required for script processing"
            )
            if deepseek_input:
                st.session_state.temp_deepseek_key = deepseek_input
                st.success("✅ DeepSeek API Key: Temporarily configured")
        
        st.divider()
        
        if st.button("🚪 Logout", type="secondary"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    
    # Main header
    st.markdown("""
    <div class="main-header">
        <h1>🎬 Film Production Master App</h1>
        <p>Standards & Practices + Production Design Analysis</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Main tabs
    tab1, tab2, tab3 = st.tabs(["📤 Upload Script", "📝 S&P Results", "🎬 Production Results"])
    
    with tab1:
        tab_upload_script()
    
    with tab2:
        tab_standards_results()
    
    with tab3:
        tab_production_results()

if __name__ == "__main__":
    main()
